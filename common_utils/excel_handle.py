# -*- coding: utf-8 -*-

# @version: v1.0
# @author : Hide
# @Project : hide
# @File : excel_handle.py
# @Software: PyCharm
# @time: 2022/7/4 11:16
# @description :


from openpyxl import load_workbook

from common_utils.api_path import data_path
from common_utils.global_vars import CaseEnum


class ReadExcel:
    def __init__(self, filename):
        self.filename = filename
        self.wb = load_workbook(self.filename)
        self.sheets = self.wb.sheetnames

    def read_excel(self, sheet=""):
        if sheet == "":
            sheets = self.sheets
        else:
            sheets = sheet

        case_data = []
        for sheet in sheets:
            wb = self.wb[sheet]
            max_row = self.wb[sheet].max_row
            for row in range(2, max_row+1):
                _dict = {}
                if wb.cell(row=row, column=CaseEnum.API_EXEC.value).value == "是":
                    _dict["id"] = wb.cell(row, column=CaseEnum.CASE_ID.value).value
                    _dict["feature"] = wb.cell(row, column=CaseEnum.CASE_FEATURE.value).value
                    _dict["title"] = wb.cell(row, column=CaseEnum.CASE_TITLE.value).value
                    _dict["url"] = wb.cell(row, column=CaseEnum.API_PATH.value).value
                    _dict["header"] = wb.cell(row, column=CaseEnum.API_HEADER.value).value
                    _dict["method"] = wb.cell(row, column=CaseEnum.API_METHOD.value).value
                    _dict["pk"] = wb.cell(row, column=CaseEnum.API_PK.value).value
                    _dict["data"] = wb.cell(row, column=CaseEnum.API_DATA.value).value
                    _dict["file"] = wb.cell(row, column=CaseEnum.API_FILE.value).value
                    _dict["extract"] = wb.cell(row, column=CaseEnum.API_EXTRACT.value).value
                    _dict["validate"] = wb.cell(row, column=CaseEnum.API_EXPECTED.value).value

                case_data.append(_dict)

        return case_data


if __name__ == '__main__':
    exec_instance = ReadExcel(data_path + "test_case.xlsx")
    result = exec_instance.read_excel()
    print(result)
    print(result[0]["head"]["uid"])






