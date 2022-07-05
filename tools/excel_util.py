# -*- coding: utf-8 -*-

# @version: v1.0
# @author : Hide
# @Project : hide
# @File : excel_util.py
# @Software: PyCharm
# @time: 2022/7/1 10:02
# @description : 读写xlsx文件

import xlrd
from openpyxl import load_workbook, Workbook


class ExcelUtil:
    def __init__(self):
        pass

    '''
    检验表头信息
    '''
    def check(self):
        pass

    def read_excel(self):
        # 1.打开要读取的excel文件
        wb = load_workbook("../data/testcase.xlsx")

        # 2.获取sheet
        sheets = wb.sheetnames
        print("sheet名:", sheets)
        sheet = wb["test"]
        print("根据指定sheet名获取sheet:", sheet)

        # 3.读取某个单元格，cell方法返回的并不是具体单元格的值，而是Cell类型
        url_address = sheet.cell(5, 2)
        print("url地址：", url_address)

        url_value = sheet.cell(3, 6).value
        print(type(url_value))

        # 4.按行遍历
        rows = []
        for row in sheet.rows:
            rows.append([cell.value for cell in row])
        print(rows)

    def create_excel(self):
        wb = Workbook()  # 创建一个workbook对象
        ws = wb.active   # ws操作sheet页
        wb.create_sheet("myfirstsheet", 0)  # 指定sheet名称
        wb.save("../data/hide.xlsx")        # 保存到指定位置

    def f(self, x):
        return x*x

    def map_demo(self):
        new_list = list(map(self.f, range(1, 7)))
        print(new_list)

    def filter_demo(self):
        filter_list = list(filter(lambda x: x % 2 == 0, range(100)))
        print(filter_list)

    def sorted_demo(self):
        new_list = ['bob', 'about', 'Zoo', 'Credit']
        sorted_list = sorted(new_list, key=str.lower, reverse=True)
        print(sorted_list)


if __name__ == '__main__':
    exec = ExcelUtil()
    exec.read_excel()
    # exec.map_demo()
    # exec.filter_demo()
    # exec.sorted_demo()
    # exec.create_excel()